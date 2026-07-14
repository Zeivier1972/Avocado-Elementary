"""Parse the FLDOE FAST item-level student data export (.xls/.xlsx).

Layout: one row per student. Columns include demographics, scale score,
achievement level, percentile, four domain "Performance" ratings, then a
repeating block of [Category, Benchmark, Points Earned, Points Possible] for
every test item. Benchmark cells look like: "NSAR|MA.3.NSO.1|MA.3.NSO.1.2".
"""
from __future__ import annotations

import io
import re

import openpyxl


def _s(v) -> str:
    return ("" if v is None else str(v)).strip()


def is_fast_export(headers: list[str]) -> bool:
    joined = " ".join(_s(h).lower() for h in headers)
    return ("scale score" in joined and "achievement level" in joined
            and "points earned" in joined)


def _level(v) -> int | None:
    m = re.search(r"(\d)", _s(v))
    return int(m.group(1)) if m else None


def _bench_code(cell: str) -> str:
    """Extract the primary benchmark code from 'NSAR|MA.3.NSO.1|MA.3.NSO.1.2'."""
    parts = [p.strip() for p in _s(cell).split("|")]
    codes = [p for p in parts if re.match(r"MA\.\w+\.[A-Z]+\.\d", p)]
    if not codes:
        return _s(cell)
    # last token is the specific benchmark; strip any "and MA...."
    return re.split(r"\s+and\s+", codes[-1])[0].strip()


def detect(data: bytes) -> bool:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        header = next(ws.iter_rows(values_only=True), ())
        wb.close()
        return is_fast_export([_s(h) for h in header])
    except Exception:
        return False


def parse_fast_export(data: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return {"students": [], "subject": "MATH", "period": ""}
    header = [_s(h) for h in rows[0]]
    hl = [h.lower() for h in header]

    def find(substr):
        for j, h in enumerate(hl):
            if substr in h:
                return j
        return None

    col_name = find("student name")
    col_district_use = find("district use")   # matches other files' IDs
    col_local = find("local id")
    col_grade = find("enrolled grade")
    col_ell = find("english language")
    col_ese = find("exceptionality")
    col_scale = find("scale score")
    col_level = find("achievement level")
    col_pct = find("percentile")
    col_reason = find("test reason")

    subject = "MATH" if "math" in " ".join(hl) else ("ELA" if "reading" in " ".join(hl) or "ela" in " ".join(hl) else "MATH")

    # domain performance columns (contain "performance", before the item block)
    first_cat = next((j for j, h in enumerate(hl) if h == "category"), None)
    domain_cols = [j for j, h in enumerate(hl)
                   if "performance" in h and (first_cat is None or j < first_cat)]

    # item blocks: every "category" column starts a [cat, bench, earned, poss] group
    cat_cols = [j for j, h in enumerate(hl) if h == "category"]

    students = []
    period = ""
    for r in rows[1:]:
        if not r or col_district_use is None:
            continue
        did = _s(r[col_district_use]).lstrip("0") or _s(r[col_local]).lstrip("0")
        if not did or not did.isdigit():
            continue
        period = period or re.sub(r"\s.*", "", _s(r[col_reason])) if col_reason is not None else "PM1"
        name = _s(r[col_name]) if col_name is not None else ""
        last, first = "", ""
        if "," in name:
            last, first = [p.strip() for p in name.split(",", 1)]
        domains = {}
        for j in domain_cols:
            dn = re.sub(r"\s*performance\s*$", "", header[j], flags=re.I)
            dn = re.sub(r"^\d+\.\s*", "", dn)
            domains[dn] = _s(r[j]) if j < len(r) else ""
        items = []
        for j in cat_cols:
            if j + 3 >= len(r):
                continue
            cat = _s(r[j])
            bench = r[j + 1]
            earned, poss = r[j + 2], r[j + 3]
            code = _bench_code(bench)
            if not code or poss in (None, "", 0):
                continue
            try:
                items.append({"category": re.sub(r"^\d+\.\s*", "", cat),
                              "benchmark_code": code,
                              "earned": float(earned or 0),
                              "possible": float(poss)})
            except (ValueError, TypeError):
                continue
        students.append({
            "district_student_id": did, "first_name": first, "last_name": last,
            "grade": _s(r[col_grade]) if col_grade is not None else "",
            "ell": _s(r[col_ell]) if col_ell is not None else "",
            "ese": _s(r[col_ese]) if col_ese is not None else "",
            "scale_score": _num(r[col_scale]) if col_scale is not None else None,
            "level": _level(r[col_level]) if col_level is not None else None,
            "percentile": _num(r[col_pct]) if col_pct is not None else None,
            "domains": domains, "items": items,
        })
    return {"students": students, "subject": subject, "period": period or "PM1"}


def _num(v):
    try:
        return float(_s(v))
    except ValueError:
        return None
