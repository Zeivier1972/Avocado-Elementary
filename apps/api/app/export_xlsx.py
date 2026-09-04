"""Excel export of the Math Goal Analysis — a teacher/coach report where the
topic-assessment average is color-coded by achievement level (L1 Red, L2 Yellow,
L3 Green, L4 Blue, L5 Orange), aligned to the FAST-based goal."""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.goal_rubric import LEVEL_COLORS

_HDR = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="38601F")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _level_fill(level):
    if not level or level not in LEVEL_COLORS:
        return None
    return PatternFill("solid", fgColor=LEVEL_COLORS[level]["hex"])


def _font_for(level):
    # yellow/orange read better with black text; others with white
    return Font(bold=True, color="000000" if level in (2, 5) else "FFFFFF")


def goal_analysis_to_xlsx(data: dict) -> bytes:
    grade = data.get("grade", "")
    wb = Workbook()
    ws = wb.active
    ws.title = f"Grade {grade} Math"

    headers = ["Student", "FAST Scale", "FAST Level", "Instructional Level",
               "Topic Goal", "Topic Avg", "Topic Level", "Status",
               "Trend", "EOY Projection"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = _HDR
        cell.fill = _HDR_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def goal_text(mn, mx):
        if mn is None:
            return ""
        return f"{mn}%" if mn == mx else f"{mn}-{mx}%"

    r = 2
    for s in data.get("students", []):
        ws.cell(r, 1, s.get("name", ""))
        ws.cell(r, 2, s.get("fast_scale"))
        # FAST level colored
        flc = ws.cell(r, 3, s.get("fast_level"))
        ffill = _level_fill(s.get("fast_level"))
        if ffill:
            flc.fill = ffill
            flc.font = _font_for(s.get("fast_level"))
        ws.cell(r, 4, s.get("instructional", ""))
        ws.cell(r, 5, goal_text(s.get("goal_min"), s.get("goal_max")))
        # Topic average colored by its achievement level (the color code)
        tav = ws.cell(r, 6, (s["topic_avg"] / 100) if s.get("topic_avg") is not None else None)
        tav.number_format = "0%"
        tlevel = s.get("topic_level")
        tfill = _level_fill(tlevel)
        if tfill:
            tav.fill = tfill
            tav.font = _font_for(tlevel)
        tl = ws.cell(r, 7, tlevel)
        if tfill:
            tl.fill = tfill
            tl.font = _font_for(tlevel)
        ws.cell(r, 8, {"above": "Above goal", "meeting": "Meeting",
                       "below": "Below goal", "no_topic": "No topic data",
                       "no_fast": "No FAST"}.get(s.get("status"), s.get("status", "")))
        ws.cell(r, 9, {"up": "↑", "down": "↓", "flat": "→"}.get(s.get("trend"), ""))
        ws.cell(r, 10, ("On track" if s.get("projected") is True
                        else "At risk" if s.get("projected") is False else ""))
        for c in range(1, 11):
            ws.cell(r, c).border = _BORDER
        r += 1

    widths = [22, 10, 10, 22, 11, 10, 11, 12, 7, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.freeze_panes = "A2"

    # Color legend
    lr = r + 1
    ws.cell(lr, 1, "Topic color code:").font = Font(bold=True)
    for i, lvl in enumerate([1, 2, 3, 4, 5]):
        cell = ws.cell(lr, 2 + i, f"L{lvl} {LEVEL_COLORS[lvl]['name']}")
        cell.fill = _level_fill(lvl)
        cell.font = _font_for(lvl)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    # Benchmark coverage on a second sheet
    cov = data.get("benchmark_coverage", [])
    if cov:
        ws2 = wb.create_sheet("Benchmark Coverage")
        ch = ["Benchmark", "Description", "Times Assessed", "Questions", "Avg %"]
        for c, h in enumerate(ch, 1):
            cell = ws2.cell(1, c, h)
            cell.font = _HDR
            cell.fill = _HDR_FILL
            cell.border = _BORDER
        for i, cvr in enumerate(cov, 2):
            ws2.cell(i, 1, cvr.get("benchmark", ""))
            ws2.cell(i, 2, cvr.get("description", ""))
            ws2.cell(i, 3, cvr.get("times_assessed"))
            ws2.cell(i, 4, cvr.get("questions"))
            ap = ws2.cell(i, 5, (cvr["avg_pct"] / 100) if cvr.get("avg_pct") is not None else None)
            ap.number_format = "0%"
        for i, w in enumerate([16, 48, 14, 11, 8], 1):
            ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w
        ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
