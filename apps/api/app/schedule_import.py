"""Parse the Avocado Master Schedule workbook into per-teacher Math and Math-DI
time blocks.

The workbook lays each grade's teachers out in day-rows (M/T/W/R/F). Row 3 holds
5-minute time columns; each subject is a horizontally-merged block. For the math
coach we extract, per teacher per day:
  - MATH blocks  → when to visit a math lesson
  - SCIENCE / SOCIAL STUDIES blocks → the window a math teacher can run Math DI
    (per the coach: DI happens during science/social-studies time).
"""
from __future__ import annotations

import io
import re
from collections import defaultdict

DAY_MAP = {"M": "Mon", "T": "Tue", "W": "Wed", "R": "Thu", "F": "Fri"}
DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri"]
# K-3 homeroom sheets the coach supports for math, including the ASD (autism
# spectrum, self-contained) K-3 classes. The parser is sheet-agnostic.
DEFAULT_SHEETS = ["Kinder & Grade 1", "Grade  2 & 3", "ASD K & 1", "ASD 2 & 3"]


def _hhmm(minutes: int) -> str:
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def _col_min_map(ws) -> dict | None:
    """Column -> minutes-from-midnight, read from whichever early row carries the
    5-minute time headers. Returns None if this sheet has no time row."""
    best_row, best_n = None, 0
    for r in range(1, 7):
        n = sum(1 for c in range(1, ws.max_column + 1)
                if hasattr(ws.cell(r, c).value, "hour"))
        if n > best_n:
            best_row, best_n = r, n
    if not best_row:
        return None
    return {c: ws.cell(best_row, c).value.hour * 60 + ws.cell(best_row, c).value.minute
            for c in range(1, ws.max_column + 1)
            if hasattr(ws.cell(best_row, c).value, "hour")}


def _classify(label: str) -> str | None:
    u = label.upper()
    if re.search(r"\bMATHEMATICS\b|\bMATH\b", u) and "CP" not in u:
        return "math"
    if "SCIENCE" in u or "SOCIAL" in u:
        return "di"
    return None


def _clean_subject(label: str) -> str:
    # drop the parenthetical co-teach / concurrency notes for a clean chip
    return re.sub(r"\s+", " ", label.split("(")[0]).strip()


def _parse_sheet(ws, canon: dict | None = None):
    # Some tabs (e.g. ASD K & 1) omit the time-header row; fall back to the
    # workbook's canonical column->time grid (all tabs share the column layout).
    col_min = _col_min_map(ws) or canon
    if not col_min:
        return []
    tmin = min(col_min)
    starts = defaultdict(list)
    for mr in ws.merged_cells.ranges:
        if mr.min_row == mr.max_row and mr.min_col >= tmin:
            starts[mr.min_row].append((mr.min_col, mr.max_col))

    def blocks(r):
        out, seen = [], set()
        for a, b in sorted(starts.get(r, [])):
            v = ws.cell(r, a).value
            if v is not None and str(v).strip():
                out.append((a, b, str(v).strip()))
                seen.update(range(a, b + 1))
        for c in sorted(col_min):
            if c in seen:
                continue
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                out.append((c, c, str(v).strip()))
        return sorted(out)

    teachers = []
    grade, program = "", ""
    r = 3
    while r <= ws.max_row:
        a = ws.cell(r, 1).value
        if a and re.search(r"grade|kinder|prek|pre-k|asd", str(a), re.I):
            grade = _norm_grade(str(a))
            program = "ASD" if re.search(r"asd", str(a), re.I) else ""
        b = ws.cell(r, 2).value
        day = ws.cell(r, 3).value
        if b and str(b).strip() and day and str(day).strip() in DAY_MAP:
            binfo = re.sub(r"\s+", " ", str(b)).strip()
            mroom = re.match(r"^([A-Z]?\d{2,3}[A-Z]?)", binfo)
            room = mroom.group(1) if mroom else ""
            name = binfo[len(room):] if room else binfo
            name = re.split(r"\(|Room|ESOL|LV|SPED|/", name)[0].strip(" -·")
            # ASD rows are room-labeled (e.g. "A07 (2nd)") with no teacher name;
            # use the room as the label (the ASD tag distinguishes it).
            if not name:
                name = room or "—"
            # a room-only ASD block may still carry the grade in its label "(2nd)"
            if not grade:
                gm = re.search(r"\((kinder|k|\d)", binfo, re.I)
                if gm:
                    grade = _norm_grade(gm.group(1))
            days = {}
            for k in range(5):
                rr = r + k
                d = ws.cell(rr, 3).value
                if not d or str(d).strip() not in DAY_MAP:
                    break
                dd = DAY_MAP[str(d).strip()]
                math, di = [], []
                for x, y, val in blocks(rr):
                    kind = _classify(val)
                    if kind == "math":
                        math.append({"start": _hhmm(col_min[x]), "end": _hhmm(col_min[y] + 5)})
                    elif kind == "di":
                        di.append({"subject": _clean_subject(val),
                                   "start": _hhmm(col_min[x]), "end": _hhmm(col_min[y] + 5)})
                days[dd] = {"math": math, "di": di}
            teachers.append({"grade": grade, "room": room, "teacher": name,
                             "program": program,
                             "teaches_math": any(days[d]["math"] for d in days),
                             "days": days})
            r += 5
        else:
            r += 1
    return teachers


def _norm_grade(raw: str) -> str:
    r = raw.strip().lower()
    if "kinder" in r:
        return "K"
    m = re.search(r"(\d+)", r)
    return m.group(1) if m else raw.strip()


def parse_master_schedule(data: bytes, sheets: list[str] | None = None) -> dict:
    """Return {'teachers': [...], 'sheets_used': [...], 'reason': str|None}."""
    try:
        import openpyxl
    except ImportError:
        return {"teachers": [], "sheets_used": [], "reason": "openpyxl not installed"}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        return {"teachers": [], "sheets_used": [],
                "reason": f"could not open workbook: {type(e).__name__}"}
    want = sheets or DEFAULT_SHEETS
    # tolerant sheet matching (trailing/double spaces vary)
    norm = {re.sub(r"\s+", " ", s).strip().lower(): s for s in wb.sheetnames}
    # Canonical column->time grid from any sheet that has one, so tabs missing
    # their time header (ASD K & 1) still resolve to clock times.
    canon = None
    for s in wb.sheetnames:
        canon = _col_min_map(wb[s])
        if canon:
            break
    teachers, used = [], []
    for target in want:
        key = re.sub(r"\s+", " ", target).strip().lower()
        real = norm.get(key)
        if real:
            teachers.extend(_parse_sheet(wb[real], canon=canon))
            used.append(real)
    if not teachers:
        return {"teachers": [], "sheets_used": used,
                "reason": "no teacher rows found on the K-3 schedule sheets"}
    return {"teachers": teachers, "sheets_used": used, "reason": None}


def to_blocks(teachers: list[dict]) -> list[dict]:
    """Flatten parsed teachers into storable rows (one per math/di block)."""
    rows = []
    for t in teachers:
        prog = t.get("program", "")
        for day, sub in t["days"].items():
            for m in sub["math"]:
                rows.append({"grade": t["grade"], "room": t["room"], "program": prog,
                             "teacher": t["teacher"], "day": day, "kind": "math",
                             "subject": "Mathematics",
                             "start": m["start"], "end": m["end"]})
            for d in sub["di"]:
                rows.append({"grade": t["grade"], "room": t["room"], "program": prog,
                             "teacher": t["teacher"], "day": day, "kind": "di",
                             "subject": d["subject"],
                             "start": d["start"], "end": d["end"]})
    return rows


def _to_min(t: str) -> int:
    h, m = t.split(":")
    return int(h) * 60 + int(m)


def _to_hhmm(x: int) -> str:
    return f"{x // 60:02d}:{x % 60:02d}"


def build_visit_plan(by_grade: dict, kind: str = "math", minutes: int = 30,
                     grade: str | None = None) -> list[dict]:
    """Greedily lay out a conflict-free week: one visit per math teacher, during
    one of their math blocks (kind='math') or DI windows (kind='di'). The coach
    is one person, so two visits never overlap on the same day. Most-constrained
    teachers (fewest options) are placed first, balancing load across days."""
    teachers = []
    for g, ts in by_grade.items():
        if grade and g != grade:
            continue
        for t in ts:
            if not t.get("teaches_math"):
                continue
            slots = []
            for day, sub in t["days"].items():
                for b in sub.get(kind, []):
                    if b.get("start") and b.get("end"):
                        slots.append({"day": day, "start": _to_min(b["start"]),
                                      "end": _to_min(b["end"]), "subject": b.get("subject", "")})
            if slots:
                teachers.append({"grade": g, "room": t["room"],
                                 "teacher": t["teacher"],
                                 "program": t.get("program", ""), "slots": slots})
    teachers.sort(key=lambda t: (len(t["slots"]), t["grade"], t["room"]))
    busy = {d: [] for d in DAY_ORDER}
    load = {d: 0 for d in DAY_ORDER}

    def overlaps(d, s, e):
        return any(s < be and bs < e for bs, be in busy[d])

    def dkey(d):
        return DAY_ORDER.index(d) if d in DAY_ORDER else 99

    plan = []
    for t in teachers:
        opts = sorted(t["slots"], key=lambda x: (load[x["day"]], dkey(x["day"]), x["start"]))
        chosen, conflict = None, False
        for o in opts:
            s, e = o["start"], min(o["start"] + minutes, o["end"])
            if not overlaps(o["day"], s, e):
                chosen = (o, s, e)
                break
        if not chosen:
            o = opts[0]
            chosen, conflict = (o, o["start"], min(o["start"] + minutes, o["end"])), True
        o, s, e = chosen
        busy[o["day"]].append((s, e))
        load[o["day"]] += 1
        plan.append({"day": o["day"], "grade": t["grade"], "room": t["room"],
                     "teacher": t["teacher"], "program": t.get("program", ""),
                     "start": _to_hhmm(s), "end": _to_hhmm(e),
                     "subject": o.get("subject", ""),
                     "block": f"{_to_hhmm(o['start'])}-{_to_hhmm(o['end'])}",
                     "conflict": conflict})
    plan.sort(key=lambda p: (dkey(p["day"]), p["start"]))
    return plan
