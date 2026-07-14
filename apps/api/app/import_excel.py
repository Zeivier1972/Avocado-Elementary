"""Parse the district's Excel workbooks (Class Lists and Math Topic Assessment
Tracker) into normalized student + assessment records.

These sheets have merged/grouped headers and vary by grade, so parsing is
tolerant: it locates the header row, uses the group-label row to disambiguate
ELA vs Math FAST columns, and derives grade from the GRADE column or sheet name.
"""
from __future__ import annotations

import io
import re

import openpyxl


def _s(v) -> str:
    return ("" if v is None else str(v)).strip()


def _num(v):
    v = _s(v).replace("%", "")
    if v in ("", "N/A", "-", "·", "="):
        return None
    try:
        return float(v)
    except ValueError:
        return None


def _norm(v) -> str:
    return re.sub(r"\s+", " ", _s(v)).lower()


def _grade_from_sheet(name: str) -> str | None:
    n = name.strip().upper()
    if n.startswith("K"):
        return "K"
    if n.startswith(("VPK", "PK", "PA", "PR", "T1")):
        return "PK"
    m = re.match(r"^(\d)\d\d", n)          # 101, 205, 301 -> 1,2,3
    if m:
        return m.group(1)
    return None


def _norm_grade(v: str, sheet_grade: str | None) -> str:
    g = _s(v).upper().replace("GRADE", "").strip()
    if g in ("0", "00", "K", "KG"):
        return "K"
    if g in ("PK", "PRE-K", "PREK"):
        return "PK"
    if g in ("1", "2", "3", "4", "5"):
        return g
    return sheet_grade or g or ""


def _period_kind(field: str):
    """Return (period, kind) from a FAST/iReady field header, else (None, None).
    kind is 'level' or 'scale'."""
    f = field
    period = None
    for p in ("pm1", "pm 1", "pm2", "pm 2", "pm3", "pm 3", "ap1", "ap 1",
              "ap2", "ap 2", "ap3", "ap 3"):
        if p in f:
            period = p.replace(" ", "").upper()
            break
    if not period and "pm3" not in f and "2025 pm3" in f:
        period = "PM3"
    if period is None:
        return None, None
    if any(k in f for k in ("dss", " ds", "sss", "scale")):
        return period, "scale"
    if any(k in f for k in ("lvl", " lv", "level")):
        return period, "level"
    return period, "level"


def _find_header(rows):
    for i, r in enumerate(rows):
        cells = [_norm(c) for c in r]
        has_id = any(c in ("id", "id number", "id#") for c in cells)
        has_name = any("last name" in c for c in cells)
        if has_id and has_name:
            return i, r
    return None, None


def parse_workbook(data: bytes) -> list[dict]:
    """Return a list of {student, assessments} dicts across all sheets."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    out: list[dict] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        hidx, header = _find_header(rows)
        if hidx is None:
            continue
        norm_hdr = [_norm(c) for c in header]
        group_row = [_norm(c) for c in rows[hidx - 1]] if hidx > 0 else []
        is_tracker = any("topic" in c for c in norm_hdr) or (
            "id number" in norm_hdr and "pm1" in norm_hdr
        )
        sheet_grade = _grade_from_sheet(sheet)
        col = {c: j for j, c in enumerate(norm_hdr)}

        def find(*names):
            for nm in names:
                if nm in col:
                    return col[nm]
            for j, c in enumerate(norm_hdr):
                if any(nm in c for nm in names):
                    return j
            return None

        id_j = find("id number", "id")
        last_j = find("last name")
        first_j = find("first name")
        grade_j = find("grade")
        ell_j = find("ell")
        ese_j = find("ese")
        if id_j is None:
            continue

        for r in rows[hidx + 1:]:
            if id_j >= len(r):
                continue
            sid = _s(r[id_j])
            if not sid or not sid.replace(".", "").isdigit():
                continue
            grade = _norm_grade(_s(r[grade_j]) if grade_j is not None else "",
                                sheet_grade)
            student = {
                "id": sid,
                "first_name": _s(r[first_j]) if first_j is not None else "",
                "last_name": _s(r[last_j]) if last_j is not None else "",
                "grade": grade,
                "flags": {},
            }
            if ell_j is not None and _s(r[ell_j]):
                student["flags"]["ell"] = _s(r[ell_j])
            if ese_j is not None and _s(r[ese_j]):
                student["flags"]["ese"] = True
                student["flags"]["ese_code"] = _s(r[ese_j])

            assessments = []
            if is_tracker:
                assessments = _tracker_assessments(norm_hdr, r)
            else:
                assessments = _classlist_assessments(norm_hdr, group_row, r)
            out.append({"student": student, "assessments": assessments,
                        "section": sheet})
    wb.close()
    return out


def _classlist_assessments(hdr, group_row, r):
    """FAST ELA/Math PM1/2/3 and iReady AP1/2 with group context."""
    res = []
    current = ""
    for j, field in enumerate(hdr):
        g = group_row[j] if j < len(group_row) else ""
        if "fast ela" in g:
            current = "FAST_ELA"
        elif "fast math" in g:
            current = "FAST_MATH"
        elif "iready" in g or "i-ready" in g:
            current = "IREADY"
        if j >= len(r):
            continue
        val = _num(r[j])
        if val is None:
            continue
        # subject from field text overrides group when explicit
        subj = "ELA" if "ela" in field else "MATH" if "math" in field else (
            "ELA" if current == "FAST_ELA" else "MATH" if current == "FAST_MATH"
            else None)
        period, kind = _period_kind(field)
        if period is None:
            continue
        if current.startswith("FAST"):
            source, subject = "FAST", subj or ("ELA" if "ELA" in current else "MATH")
        elif current == "IREADY":
            source, subject = "IREADY", subj
        else:
            continue
        if subject is None:
            continue
        rec = {"source": source, "subject": subject, "period": period,
               "level": None, "scale_score": None, "label": field}
        if kind == "scale":
            rec["scale_score"] = val
        else:
            rec["level"] = val
        res.append(rec)
    return res


def _tracker_assessments(hdr, r):
    """Math tracker: PM1/2/3 = FAST Math level, AP1/2 = iReady Math level,
    Topic N = topic percent."""
    res = []
    for j, field in enumerate(hdr):
        if j >= len(r):
            continue
        val = _num(r[j])
        if val is None:
            continue
        if field in ("pm1", "pm2", "pm3"):
            res.append({"source": "FAST", "subject": "MATH",
                        "period": field.upper(), "level": val,
                        "scale_score": None, "label": field})
        elif field in ("ap1", "ap2", "ap3"):
            res.append({"source": "IREADY", "subject": "MATH",
                        "period": field.upper(), "level": val,
                        "scale_score": None, "label": field})
        elif field.startswith("topic"):
            m = re.search(r"topic\s*([\d/]+)", field)
            if m and 0 <= val <= 1.5:
                res.append({"source": "TOPIC", "subject": "MATH",
                            "period": "TP" + m.group(1).replace("/", "_"),
                            "percent": val, "label": field})
    return res
